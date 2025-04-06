from flask import Flask, render_template, request, send_file
import os
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side

app = Flask(__name__, template_folder='src')

os.makedirs('excel_exports', exist_ok=True)

classes = ["10 ជ"]
students = [
    "ខាន់ សិរីណុង", "ខឿន សារ៉ា", "ខៀវ បញ្ញាពេជ្រ", "គ្រីយ៉ា ណូរ៉ា",
    "ឆើត ស្រីអូន", "ជាន សុធារិទ្ធ", "ឈិន ឡុងផាត់", "ឌឿន ស្រីវី",
    "ណាន វណ្ណៈ", "ណារ៉ាត់ ដាវីត",  "ណែន ស្រីណុន", "ប៊ន ឌីណា", "ផន ស្រីពេជ្រ",
    "ផល សុភ័ក្រទី", "ផល សុភ័ស", "ផល្លី ឡៃហ៊ួ", "ផាន់នី ស្រីម៉ៅ",
    "ភាព សុផាត", "ម៉ម សុផាន់ណា", "ម៉ី សុខឡៃ", "ម៉ុង រ៉ុង",
    "មាស សុវណ្ណវិជ្ជរ៉ា", "មឿន រីណា", "យ៉ម ចាន់ម៉ីលី", "រ៉ន សេងថៃ",
    "រិត វណ្ណារី", "រិទ្ធ ផាវីន", "រុំ សុខខេង", "លី ស្រីពីន",
    "វ៉ា កាន្នីដា", "វ៉ាន់ធី ដាវីត", "រ៉ូ ធាយុ", "វី រត្ននា",
    "វៃ ហុង", "សន សុវណ្ណា", "សារ៉ុន ធារី", "សេង លឹមហេង",
    "ហន តុលាកា", "ហាន សុភ័ក្រ", "ហេង ម៉េងហ៊ី", "ហៃ វីហាក់",
    "ឡង ដាលីន", "ឯម ម៉េងហ៊ាង", "ឯល សាហាទី", "អុី ប៊ុនហេង",
    "អុិច អុី", "អាត វីរៈ", "អូន មករា", "អូន ម៉េង",
    "អេន ភីរ៉ាត់", "ឆុន ថុនា", "យីម រតនៈ"
]

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        date = request.form.get("date")
        start_time = request.form.get("start_time")
        end_time = request.form.get("end_time")
        subject = request.form.get("subject")
        class_name = request.form.get("class", classes[0])
        action = request.form.get("action")
        selected_students = request.form.getlist("selected_students")

        if not selected_students:
            return render_template("index.html", classes=classes, students=students,
                                   error="Please select at least one student")

        attendances = []
        for student in selected_students:
            attendance_key = f"attendance_{student}"

            attendance_value = request.form.get(attendance_key, "A")
            attendances.append(attendance_value)

        print(f"Selected students: {selected_students}")
        print(f"Attendance values: {attendances}")

        if action == "export_excel":
            try:
                excel_filename = export_to_excel(selected_students, attendances, date, f"{start_time} - {end_time}", subject, class_name)
                return send_file(excel_filename, as_attachment=True, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                return render_template("index.html", classes=classes, students=students,
                                       error=f"Error generating Excel file: {str(e)}")

        return render_template("attendance.html", class_name=class_name, date=date, time_range=f"{start_time} - {end_time}",
                               subject=subject, students=zip(selected_students, attendances))

    return render_template("index.html", classes=classes, students=students)


def export_to_excel(students, attendances, date, time_range, subject, class_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    header_font = Font(name='Arial', size=14, bold=True)
    normal_font = Font(name='Arial', size=10)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    ws.append(["វិទ្យាល័យហ៊ុនសែនស្នួល", "", "", "Honsen Snoul High School"])
    ws.append([f"Attendance Record - Class {class_name}", "", "", ""])
    ws.append(["Date:", date, "Time:", time_range])
    ws.append(["Subject:", subject, "", ""])
    ws.append([""])

    headers = ["No.", "Student Name", "Present", "Absent"]
    ws.append(headers)
    for col in range(1, 5):
        cell = ws.cell(row=6, column=col)
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 8  
    ws.column_dimensions['B'].width = 40  
    ws.column_dimensions['C'].width = 15 
    ws.column_dimensions['D'].width = 15  

    for i, (student, attendance) in enumerate(zip(students, attendances), start=1):
        row = [i, student, "✓" if attendance == "P" else "", "✓" if attendance == "A" else ""]
        ws.append(row)

        for col in range(1, 5):
            cell = ws.cell(row=i + 6, column=col)
            cell.border = thin_border
            cell.font = normal_font
            if col == 2:  
                cell.alignment = Alignment(horizontal='left')
            else:  
                cell.alignment = Alignment(horizontal='center')

    for row in range(1, 5):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='center')

    filename = f"excel_exports/Attendance_{class_name.replace(' ', '_')}_{subject.replace(' ', '_')}_{date.replace('-', '')}.xlsx"
    wb.save(filename)
    return filename


if __name__ == "__main__":
    app.run(debug=True, host='0.0.0.0', port=8080)